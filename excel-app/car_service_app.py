import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
import json
import os
from datetime import datetime, timedelta

class CarServiceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Vehicle Maintenance Manager")
        self.root.geometry("1200x700")
        
        # Initialize themes
        self.themes = {
            "Default": {
                "background": "#ffffff",
                "foreground": "#000000",
                "selected": "#0078D7",
                "header_bg": "#f0f0f0",
                "odd_row": "#f5f5f5",
                "even_row": "#ffffff",
                "button_bg": "#e1e1e1",
                "button_fg": "#000000",
                "status_bg": "#f8f8f8",
                "accent": "#0078D7",
                "date_new": "#28a745",
                "date_recent": "#ffc107",
                "date_old": "#dc3545",
                "column_date": "#4a69bd",
                "column_cost": "#6ab04c",
                "column_status": "#fa8231",
                "column_default": "#2f3640"
            },
            "Dark Mode": {
                "background": "#2d2d2d",
                "foreground": "#ffffff",
                "selected": "#0078D7",
                "header_bg": "#1e1e1e",
                "odd_row": "#333333",
                "even_row": "#2d2d2d",
                "button_bg": "#404040",
                "button_fg": "#ffffff",
                "status_bg": "#1e1e1e",
                "accent": "#0078D7",
                "date_new": "#28a745",
                "date_recent": "#ffc107",
                "date_old": "#dc3545",
                "column_date": "#4a69bd",
                "column_cost": "#6ab04c",
                "column_status": "#fa8231",
                "column_default": "#ffffff"
            },
            "Forest": {
                "background": "#f5f9f5",
                "foreground": "#2c4a2c",
                "selected": "#4caf50",
                "header_bg": "#e8f5e9",
                "odd_row": "#f1f8f1",
                "even_row": "#ffffff",
                "button_bg": "#81c784",
                "button_fg": "#ffffff",
                "status_bg": "#e8f5e9",
                "accent": "#4caf50",
                "date_new": "#28a745",
                "date_recent": "#ffc107",
                "date_old": "#dc3545",
                "column_date": "#4a69bd",
                "column_cost": "#6ab04c",
                "column_status": "#fa8231",
                "column_default": "#2f3640"
            },
            "Ocean": {
                "background": "#f5f9fc",
                "foreground": "#1a237e",
                "selected": "#1976d2",
                "header_bg": "#e3f2fd",
                "odd_row": "#f1f8fc",
                "even_row": "#ffffff",
                "button_bg": "#64b5f6",
                "button_fg": "#ffffff",
                "status_bg": "#e3f2fd",
                "accent": "#1976d2",
                "date_new": "#28a745",
                "date_recent": "#ffc107",
                "date_old": "#dc3545",
                "column_date": "#4a69bd",
                "column_cost": "#6ab04c",
                "column_status": "#fa8231",
                "column_default": "#2f3640"
            },
            "Professional": {
                "background": "#ffffff",
                "foreground": "#2c3e50",
                "selected": "#34495e",
                "header_bg": "#ecf0f1",
                "odd_row": "#f9f9f9",
                "even_row": "#ffffff",
                "button_bg": "#bdc3c7",
                "button_fg": "#2c3e50",
                "status_bg": "#ecf0f1",
                "accent": "#34495e",
                "date_new": "#28a745",
                "date_recent": "#ffc107",
                "date_old": "#dc3545",
                "column_date": "#4a69bd",
                "column_cost": "#6ab04c",
                "column_status": "#fa8231",
                "column_default": "#2f3640"
            }
        }
        
        # Load saved theme or use default
        self.current_theme = self.load_theme()
        
        # Add column type tracking
        self.date_column_index = None
        self.cost_column_index = None
        self.status_columns = []
        
        # Initialize main frame
        self.main_frame = ttk.Frame(self.root, padding="10")
        self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        self.main_frame.grid_rowconfigure(2, weight=1)
        self.main_frame.grid_columnconfigure(0, weight=1)
        
        # Initialize variables
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.search_records)
        
        # Initialize Excel file
        self.excel_file = None
        self.headers = []
        self.data = []
        
        # Select Excel file
        self.select_excel_file()
        if not self.excel_file:
            self.root.destroy()
            return
            
        # Load Excel data
        self.load_data()
        
        # Create theme selector and search bar
        self.create_theme_selector()
        
        # Create status bar first
        self.create_status_bar()
        
        # Create Treeview
        self.create_treeview()
        
        # Create buttons
        self.create_buttons()
        
        # Update total cost
        self.update_total_cost()
        
        # Apply current theme
        self.apply_theme(self.current_theme)

    def load_theme(self):
        try:
            with open('theme_preference.json', 'r') as f:
                return json.load(f)['theme']
        except:
            return "Default"
    
    def save_theme(self, theme_name):
        with open('theme_preference.json', 'w') as f:
            json.dump({'theme': theme_name}, f)
    
    def create_theme_selector(self):
        # Create a frame for the top bar that will contain search and theme
        top_bar = ttk.Frame(self.main_frame)
        top_bar.grid(row=0, column=0, columnspan=4, pady=5, sticky=(tk.W, tk.E))
        
        # Create theme frame on the right side
        theme_frame = ttk.LabelFrame(top_bar, text="Theme Settings", padding="5")
        theme_frame.pack(side=tk.RIGHT, padx=10)
        
        # Add theme selector
        self.theme_var = tk.StringVar(value=self.current_theme)
        theme_menu = ttk.OptionMenu(
            theme_frame,
            self.theme_var,
            self.current_theme,
            *self.themes.keys(),
            command=self.change_theme
        )
        theme_menu.pack(side=tk.LEFT, padx=5)
        
        # Adjust the search frame to be in the top bar
        search_frame = ttk.LabelFrame(top_bar, text="Search", padding="5")
        search_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=10)
        
        self.search_var = tk.StringVar()
        self.search_var.trace('w', self.search_records)
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
    
    def change_theme(self, theme_name):
        self.current_theme = theme_name
        self.apply_theme(theme_name)
        self.save_theme(theme_name)
    
    def apply_theme(self, theme_name):
        """Apply the selected theme to all widgets"""
        theme = self.themes[theme_name]
        style = ttk.Style()
        
        # Configure styles
        style.configure("TFrame", background=theme["background"])
        style.configure("TLabel", background=theme["background"], foreground=theme["foreground"])
        style.configure("TButton", background=theme["button_bg"], foreground=theme["button_fg"])
        style.configure("Treeview", background=theme["background"], foreground=theme["foreground"], fieldbackground=theme["background"])
        style.configure("Treeview.Heading", background=theme["header_bg"], foreground=theme["foreground"])
        style.map("Treeview", background=[("selected", theme["selected"])], foreground=[("selected", "white")])
        
        # Configure tree colors
        self.tree.tag_configure("oddrow", background=theme["odd_row"])
        self.tree.tag_configure("evenrow", background=theme["even_row"])
        
        # Configure date tags
        self.tree.tag_configure("date_new", foreground=theme["date_new"])
        self.tree.tag_configure("date_recent", foreground=theme["date_recent"])
        self.tree.tag_configure("date_old", foreground=theme["date_old"])
        
        # Configure column type tags
        self.tree.tag_configure("col_date", foreground=theme["column_date"])
        self.tree.tag_configure("col_cost", foreground=theme["column_cost"])
        self.tree.tag_configure("col_status", foreground=theme["column_status"])
        
        # Save theme preference
        with open('theme_preference.json', 'w') as f:
            json.dump({'theme': theme_name}, f)
        
        self.current_theme = theme_name
        
        # Update all widgets with new theme
        self.root.configure(bg=theme["background"])
        self.main_frame.configure(style="TFrame")
        
        # Refresh treeview to apply new colors
        self.refresh_treeview()
    
    def select_excel_file(self):
        """Open file dialog to select Excel file"""
        initial_dir = os.path.dirname(os.path.abspath(__file__))
        self.excel_file = filedialog.askopenfilename(
            initialdir=initial_dir,
            title="Select Excel File",
            filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*"))
        )
    
    def load_data(self):
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            sheet = wb.active
            
            # Get headers from the first row
            self.headers = [cell.value if cell.value is not None else "" for cell in sheet[1]]
            
            # Identify column types
            self.identify_column_types()
            
            # Get data from remaining rows
            self.data = []
            for row in sheet.iter_rows(min_row=2):
                row_data = [cell.value if cell.value is not None else "" for cell in row]
                self.data.append(row_data)
            
            wb.close()
        except Exception as e:
            messagebox.showerror("Error", f"Error loading Excel file: {str(e)}")
            self.root.destroy()
    
    def identify_column_types(self):
        """Identify the type of each column based on header names"""
        self.date_column_index = None
        self.cost_column_index = None
        self.status_columns = []
        
        date_keywords = ['date', 'day', 'when']
        cost_keywords = ['cost', 'price', 'amount', 'total', 'charge', 'fee']
        status_keywords = ['status', 'condition', 'maintenance', 'service', 'repair', 'mileage', 'odometer']
        
        for i, header in enumerate(self.headers):
            header_lower = str(header).lower()
            
            # Check for date column
            if any(keyword in header_lower for keyword in date_keywords):
                self.date_column_index = i
            
            # Check for cost column
            elif any(keyword in header_lower for keyword in cost_keywords):
                self.cost_column_index = i
            
            # Check for status/maintenance columns
            elif any(keyword in header_lower for keyword in status_keywords):
                self.status_columns.append(i)

    def get_column_color(self, column_index):
        """Get the color for a specific column based on its type"""
        if column_index == self.date_column_index:
            return self.themes[self.current_theme]['column_date']
        elif column_index == self.cost_column_index:
            return self.themes[self.current_theme]['column_cost']
        elif column_index in self.status_columns:
            return self.themes[self.current_theme]['column_status']
        else:
            return self.themes[self.current_theme]['column_default']

    def create_status_bar(self):
        """Create status bar with three sections: records, costs, and legends"""
        self.status_frame = ttk.Frame(self.main_frame)
        self.status_frame.grid(row=4, column=0, columnspan=4, sticky=(tk.W, tk.E), pady=(5, 0))
        
        # Left section - Record counts
        records_frame = ttk.LabelFrame(self.status_frame, text="Records")
        records_frame.pack(side=tk.LEFT, padx=5, fill=tk.Y)
        
        self.status_label = ttk.Label(records_frame, text="Total records: 0")
        self.status_label.pack(side=tk.TOP, padx=5, pady=2)
        
        self.record_count_label = ttk.Label(records_frame, text="Showing: 0")
        self.record_count_label.pack(side=tk.TOP, padx=5, pady=2)
        
        # Middle section - Costs
        costs_frame = ttk.LabelFrame(self.status_frame, text="Costs")
        costs_frame.pack(side=tk.LEFT, padx=5, fill=tk.Y)
        
        self.total_cost_label = ttk.Label(costs_frame, text="Total: $0.00")
        self.total_cost_label.pack(side=tk.TOP, padx=5, pady=2)
        
        self.filtered_cost_label = ttk.Label(costs_frame, text="Filtered: $0.00")
        self.filtered_cost_label.pack(side=tk.TOP, padx=5, pady=2)
        
        # Right section - Legends
        legend_frame = ttk.Frame(self.status_frame)
        legend_frame.pack(side=tk.RIGHT, padx=5)
        
        # Date colors legend
        date_frame = ttk.LabelFrame(legend_frame, text="Service Dates")
        date_frame.pack(side=tk.LEFT, padx=5, fill=tk.BOTH)
        
        date_grid = ttk.Frame(date_frame)
        date_grid.pack(padx=5, pady=2)
        
        # New dates
        ttk.Label(date_grid, text="■", foreground=self.themes[self.current_theme]['date_new']).grid(row=0, column=0, padx=2)
        ttk.Label(date_grid, text="< 3 months", foreground=self.themes[self.current_theme]['date_new']).grid(row=0, column=1, sticky=tk.W)
        
        # Recent dates
        ttk.Label(date_grid, text="■", foreground=self.themes[self.current_theme]['date_recent']).grid(row=1, column=0, padx=2)
        ttk.Label(date_grid, text="3-6 months", foreground=self.themes[self.current_theme]['date_recent']).grid(row=1, column=1, sticky=tk.W)
        
        # Old dates
        ttk.Label(date_grid, text="■", foreground=self.themes[self.current_theme]['date_old']).grid(row=2, column=0, padx=2)
        ttk.Label(date_grid, text="> 6 months", foreground=self.themes[self.current_theme]['date_old']).grid(row=2, column=1, sticky=tk.W)
        
        # Column type legend
        col_frame = ttk.LabelFrame(legend_frame, text="Column Types")
        col_frame.pack(side=tk.LEFT, padx=5, fill=tk.BOTH)
        
        col_grid = ttk.Frame(col_frame)
        col_grid.pack(padx=5, pady=2)
        
        # Date columns
        ttk.Label(col_grid, text="■", foreground=self.themes[self.current_theme]['column_date']).grid(row=0, column=0, padx=2)
        ttk.Label(col_grid, text="Dates", foreground=self.themes[self.current_theme]['column_date']).grid(row=0, column=1, sticky=tk.W)
        
        # Cost columns
        ttk.Label(col_grid, text="■", foreground=self.themes[self.current_theme]['column_cost']).grid(row=1, column=0, padx=2)
        ttk.Label(col_grid, text="Costs", foreground=self.themes[self.current_theme]['column_cost']).grid(row=1, column=1, sticky=tk.W)
        
        # Status columns
        ttk.Label(col_grid, text="■", foreground=self.themes[self.current_theme]['column_status']).grid(row=2, column=0, padx=2)
        ttk.Label(col_grid, text="Status", foreground=self.themes[self.current_theme]['column_status']).grid(row=2, column=1, sticky=tk.W)

    def update_record_count(self):
        """Update record counts in status bar"""
        visible_records = len(self.tree.get_children())
        total_records = len(self.data)
        self.status_label.config(text=f"Total records: {total_records}")
        if self.search_var.get():
            self.record_count_label.config(text=f"Showing: {visible_records}")
        else:
            self.record_count_label.config(text=f"Showing: {total_records}")

    def update_total_cost(self):
        """Update cost information in status bar"""
        if self.cost_column_index is None:
            return
            
        # Calculate total cost for all records
        total_cost = self.calculate_total_cost()
        self.total_cost_label.config(text=f"Total: ${total_cost:,.2f}")
        
        # Calculate total cost for filtered/visible records
        visible_items = []
        for item_id in self.tree.get_children():
            visible_items.append(self.tree.item(item_id)['values'])
        filtered_cost = self.calculate_total_cost(visible_items)
        self.filtered_cost_label.config(text=f"Filtered: ${filtered_cost:,.2f}")
        
        # Update record count
        self.update_record_count()
    
    def create_treeview(self):
        tree_frame = ttk.Frame(self.main_frame)
        tree_frame.grid(row=1, column=0, columnspan=4, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Create Treeview with striped rows
        self.tree = ttk.Treeview(tree_frame, show='headings', style="Striped.Treeview")
        
        # Configure style
        style = ttk.Style()
        style.configure("Striped.Treeview", background="white")
        style.configure("Striped.Treeview.Heading", font=('Arial', 10, 'bold'))
        style.map("Striped.Treeview", background=[('selected', '#0078D7')])
        
        # Configure columns
        self.tree["columns"] = self.headers
        
        # Create scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        # Grid layout
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        vsb.grid(row=0, column=1, sticky=(tk.N, tk.S))
        hsb.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        # Configure column headings and widths
        for col in self.headers:
            self.tree.heading(col, text=col)
            # Calculate max width
            max_width = max(
                len(str(col)),
                max((len(str(row[self.headers.index(col)])) if row[self.headers.index(col)] else 0) for row in self.data) if self.data else 0
            ) * 10 + 10
            self.tree.column(col, width=min(max_width, 200))
        
        # Configure grid weights
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        self.refresh_treeview()
        self.tree.bind('<Double-1>', self.on_double_click)
    
    def refresh_treeview(self, search_term=''):
        """Refresh the treeview with filtered data if search term is provided"""
        # Clear the treeview
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        # Add data to treeview with optional filtering
        for row in self.data:
            if search_term:
                # Check if search term is in any column
                if not any(str(value).lower().find(search_term.lower()) >= 0 for value in row):
                    continue
            
            values = []
            for col_idx, value in enumerate(row):
                if isinstance(value, datetime):  
                    values.append(value.strftime('%Y-%m-%d'))
                else:
                    values.append(value)
            
            self.tree.insert('', tk.END, values=values)
        
        # Update status bar
        self.update_record_count()
        self.update_total_cost()
    
    def create_buttons(self):
        button_frame = ttk.Frame(self.main_frame)
        button_frame.grid(row=2, column=0, columnspan=4, pady=10)
        
        ttk.Button(button_frame, text="Add Record", command=self.add_record).grid(row=0, column=0, padx=5)
        ttk.Button(button_frame, text="Edit Record", command=self.edit_record).grid(row=0, column=1, padx=5)
        ttk.Button(button_frame, text="Delete Record", command=self.delete_record).grid(row=0, column=2, padx=5)
        ttk.Button(button_frame, text="Save Changes", command=self.save_changes).grid(row=0, column=3, padx=5)
        ttk.Button(button_frame, text="Refresh", command=lambda: self.refresh_treeview()).grid(row=0, column=4, padx=5)
    
    def create_edit_window(self, title, values=None):
        edit_window = tk.Toplevel(self.root)
        edit_window.title(title)
        edit_window.geometry("600x400")
        
        form_frame = ttk.Frame(edit_window, padding="10")
        form_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        entries = {}
        for i, col in enumerate(self.headers):
            ttk.Label(form_frame, text=f"{col}:").grid(row=i, column=0, padx=5, pady=5, sticky=tk.W)
            var = tk.StringVar(value=str(values[i]) if values and values[i] is not None else '')
            entry = ttk.Entry(form_frame, textvariable=var)
            entry.grid(row=i, column=1, padx=5, pady=5, sticky=(tk.W, tk.E))
            entries[col] = var
        
        def save_record():
            new_values = [entries[col].get() for col in self.headers]
            
            if values:  # Editing existing record
                idx = self.tree.selection()[0]
                item_values = self.tree.item(idx)['values']
                for i, row in enumerate(self.data):
                    if all(str(a) == str(b) for a, b in zip(row, item_values)):
                        self.data[i] = new_values
                        break
            else:  # Adding new record
                self.data.append(new_values)
            
            self.refresh_treeview()
            edit_window.destroy()
        
        ttk.Button(form_frame, text="Save", command=save_record).grid(row=len(self.headers), column=0, columnspan=2, pady=20)
    
    def add_record(self):
        self.create_edit_window("Add New Record")
    
    def edit_record(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select a record to edit")
            return
        
        values = self.tree.item(selected_item)['values']
        self.create_edit_window("Edit Record", values)
    
    def on_double_click(self, event):
        item = self.tree.selection()
        if item:
            self.edit_record()
    
    def delete_record(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Warning", "Please select a record to delete")
            return
        
        if messagebox.askyesno("Confirm", "Are you sure you want to delete this record?"):
            item_values = self.tree.item(selected_item)['values']
            for i, row in enumerate(self.data):
                if all(str(a) == str(b) for a, b in zip(row, item_values)):
                    del self.data[i]
                    break
            self.refresh_treeview()
    
    def save_changes(self):
        try:
            # Clear the worksheet and write headers
            self.ws.delete_rows(1, self.ws.max_row)
            for col, header in enumerate(self.headers, 1):
                self.ws.cell(row=1, column=col, value=header)
            
            # Write data
            for row_idx, row_data in enumerate(self.data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    self.ws.cell(row=row_idx, column=col_idx, value=value)
            
            self.wb.save(self.excel_file)
            messagebox.showinfo("Success", "Changes saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Error saving changes: {str(e)}")
    
    def search_records(self, *args):
        search_term = self.search_var.get().lower()
        self.refresh_treeview(search_term)

    def calculate_total_cost(self, items=None):
        if self.cost_column_index is None:
            return 0.0
            
        total = 0.0
        items = items or self.data
        
        for row in items:
            try:
                cost_str = str(row[self.cost_column_index])
                # Remove currency symbols and commas
                cost_str = ''.join(c for c in cost_str if c.isdigit() or c in '.-')
                cost = float(cost_str)
                total += cost
            except (ValueError, IndexError):
                continue
                
        return total

    def get_date_tag(self, date_str):
        try:
            from datetime import datetime, timedelta
            
            # Try to parse the date string
            try:
                date = datetime.strptime(str(date_str), '%Y-%m-%d')
            except ValueError:
                try:
                    date = datetime.strptime(str(date_str), '%m/%d/%Y')
                except ValueError:
                    return 'default'  # Return default if date can't be parsed
            
            today = datetime.now()
            three_months_ago = today - timedelta(days=90)  
            six_months_ago = today - timedelta(days=180)   
            
            if date >= three_months_ago:
                return 'date_new'          
            elif date >= six_months_ago:
                return 'date_recent'       
            else:
                return 'date_old'          
        except:
            return 'default'

if __name__ == "__main__":
    root = tk.Tk()
    app = CarServiceApp(root)
    root.mainloop()
